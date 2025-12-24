"""
Hotel Gap Analysis Dashboard

Run with: streamlit run app.py
"""

import streamlit as st
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
import polars as pl
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, timedelta
from pathlib import Path
import io

from db import get_hotels, get_suppliers, get_hotel_rates, get_meal_types, get_room_types_by_hotel, test_connection
from gap_analyzer import (
    rates_to_dataframe,
    expand_date_ranges,
    generate_all_hotel_gaps,
    get_supplier_summary,
    prepare_csv_export_template,
    BOARD_EQUIVALENTS,
    REQUIRED_OCCUPANCIES,
    OCCUPANCY_CODES,
    BOARD_TO_MEAL_CODE,
)
from graphql_client import insert_hotel_rate, test_hasura_connection

# Page config
st.set_page_config(
    page_title="Hotel Gap Analysis",
    page_icon="🏨",
    layout="wide",
    initial_sidebar_state="expanded",
)


# Cached data loaders - fetch once, filter client-side
@st.cache_data(show_spinner=False)
def load_all_hotels():
    """Load all hotels from DB (cached)."""
    return get_hotels()


@st.cache_data(show_spinner=False)
def load_all_suppliers():
    """Load all suppliers from DB (cached)."""
    return get_suppliers()


@st.cache_data(show_spinner=False)
def load_all_rates():
    """Load all rates from DB (cached). Returns Polars DataFrame."""
    rates = get_hotel_rates()
    if not rates:
        return pl.DataFrame()
    return rates_to_dataframe(rates)


@st.cache_data(show_spinner=False)
def load_all_meal_types():
    """Load all meal types from DB (cached)."""
    return get_meal_types()


@st.cache_data(show_spinner=False)
def load_room_types_for_hotel(hotel_id: str):
    """Load room types for a specific hotel (cached)."""
    return get_room_types_by_hotel(hotel_id)


def load_auth_config():
    """Load authentication config from config.yaml or environment variables."""
    import os

    # Try loading from environment variables first (for Railway/production)
    if os.getenv("AUTH_USERNAME") and os.getenv("AUTH_PASSWORD_HASH"):
        return {
            "credentials": {
                "usernames": {
                    os.getenv("AUTH_USERNAME", "product_team"): {
                        "email": os.getenv("AUTH_EMAIL", "product@tripon.com"),
                        "name": os.getenv("AUTH_NAME", "Product Team"),
                        "password": os.getenv("AUTH_PASSWORD_HASH"),
                    }
                }
            },
            "cookie": {
                "expiry_days": int(os.getenv("AUTH_COOKIE_EXPIRY", "30")),
                "key": os.getenv("AUTH_COOKIE_KEY", "gap_analysis_secret_key"),
                "name": os.getenv("AUTH_COOKIE_NAME", "gap_analysis_cookie"),
            }
        }

    # Fall back to config.yaml (for local development)
    config_path = Path(__file__).parent / "config.yaml"
    if not config_path.exists():
        st.error("config.yaml not found. Please create it from config.yaml.example or set AUTH_* environment variables")
        st.stop()

    with open(config_path) as file:
        return yaml.load(file, Loader=SafeLoader)


def main_dashboard():
    """Main dashboard content (shown after authentication)."""
    st.title("🏨 Hotel Gap Analysis")

    # Test database connection
    if not test_connection():
        st.error("Cannot connect to database. Please check DATABASE_URL in .env file.")
        st.stop()

    # Sidebar configuration
    st.sidebar.header("Configuration")

    # Refresh data button
    st.sidebar.subheader("Data")
    if st.sidebar.button("🔄 Refresh Data", help="Clear cache and reload from database"):
        load_all_hotels.clear()
        load_all_suppliers.clear()
        load_all_rates.clear()
        st.rerun()

    # Load cached data
    with st.spinner("Loading data from database..."):
        all_hotels = load_all_hotels()
        all_suppliers = load_all_suppliers()
        all_rates_df = load_all_rates()

    if len(all_rates_df) == 0:
        st.warning("No rates found in database.")
        return

    st.sidebar.success("Data cached")

    # Date range
    st.sidebar.subheader("Analysis Period")
    start_date = st.sidebar.date_input(
        "Start Date", date.today(), key="analysis_start"
    )
    end_date = st.sidebar.date_input(
        "End Date", date.today() + timedelta(days=365), key="analysis_end"
    )

    # City filter
    st.sidebar.subheader("Filters")
    city_filter = st.sidebar.selectbox(
        "City", ["All", "Makkah", "Madinah"], key="city_filter"
    )

    # Star rating filter
    star_options = ["All", 1, 2, 3, 4, 5]
    star_filter = st.sidebar.selectbox(
        "Star Rating", star_options, key="star_filter"
    )

    # Build supplier options from cached data
    supplier_options = {"All": None}
    supplier_options.update({s["name"]: s["id"] for s in all_suppliers})
    supplier_filter = st.sidebar.selectbox(
        "Supplier", list(supplier_options.keys()), key="supplier_filter"
    )

    # Build hotel options - filter by city if selected
    filtered_hotels = all_hotels
    if city_filter != "All":
        filtered_hotels = [h for h in all_hotels if h["city"] == city_filter]

    hotel_options = {"All": None}
    hotel_options.update({f"{h['hotel_name']} ({h['city']})": h["hotel_id"] for h in filtered_hotels})
    hotel_filter = st.sidebar.selectbox(
        "Hotel", list(hotel_options.keys()), key="hotel_filter"
    )

    # Apply filters client-side on cached DataFrame
    df = all_rates_df

    # Date filter: rates that overlap with analysis period
    df = df.filter(
        (pl.col("end_date") >= start_date) & (pl.col("start_date") <= end_date)
    )

    # City filter
    if city_filter != "All":
        df = df.filter(pl.col("city") == city_filter)

    # Star rating filter
    if star_filter != "All":
        df = df.filter(pl.col("star_rating") == star_filter)

    # Supplier filter
    if supplier_options[supplier_filter] is not None:
        df = df.filter(pl.col("supplier_id") == supplier_options[supplier_filter])

    # Hotel filter
    if hotel_options[hotel_filter] is not None:
        df = df.filter(pl.col("hotel_id") == hotel_options[hotel_filter])

    if len(df) == 0:
        st.warning("No rates found for the selected filters.")
        return

    daily_df = expand_date_ranges(df)

    # Sidebar stats
    st.sidebar.markdown("---")
    st.sidebar.caption(f"Rates loaded: {len(df):,}")
    st.sidebar.caption(f"Hotels: {df['hotel_id'].n_unique()}")
    st.sidebar.caption(f"Daily records: {len(daily_df):,}")

    # Tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📋 Gap Report", "👥 By Supplier", "📊 Summary", "📅 Visualizations", "✏️ Fill Gaps"])

    # TAB 1: Gap Report
    with tab1:
        st.subheader("Gap Analysis Configuration")

        # Exclusion periods
        st.markdown("#### 📅 Excluded Periods")
        st.caption("Add date ranges to exclude from gap analysis (e.g., Ramadan, Hajj)")

        if "exclusions" not in st.session_state:
            st.session_state.exclusions = []

        with st.expander("Add Exclusion Period", expanded=len(st.session_state.exclusions) == 0):
            col1, col2, col3 = st.columns([2, 2, 3])
            with col1:
                excl_start = st.date_input("Exclusion Start", key="excl_start")
            with col2:
                excl_end = st.date_input("Exclusion End", key="excl_end")
            with col3:
                excl_reason = st.text_input("Reason (optional)", key="excl_reason")

            if st.button("Add Exclusion"):
                if excl_start <= excl_end:
                    st.session_state.exclusions.append({
                        "start": excl_start,
                        "end": excl_end,
                        "reason": excl_reason or "No reason",
                    })
                    st.rerun()

        if st.session_state.exclusions:
            st.markdown("**Current Exclusions:**")
            for i, excl in enumerate(st.session_state.exclusions):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(f"• {excl['start'].strftime('%d-%m-%Y')} to {excl['end'].strftime('%d-%m-%Y')} - {excl['reason']}")
                with col2:
                    if st.button("🗑️", key=f"del_excl_{i}"):
                        st.session_state.exclusions.pop(i)
                        st.rerun()

        st.markdown("---")

        # Requirements
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 🍽️ Required Boards")
            required_boards = st.multiselect(
                "Select boards to check",
                options=list(BOARD_EQUIVALENTS.keys()),
                default=["Room Only", "Breakfast"],
                key="required_boards",
            )

        with col2:
            st.markdown("#### 🛏️ Required Occupancies")
            required_occupancies = st.multiselect(
                "Select occupancies to check",
                options=list(REQUIRED_OCCUPANCIES.keys()),
                default=["Double", "Triple", "Quad"],
                key="required_occupancies",
            )

        st.markdown("---")

        # Generate report
        if st.button("🔍 Generate Gap Report", type="primary"):
            with st.spinner("Analyzing gaps..."):
                gaps_df = generate_all_hotel_gaps(
                    daily_df,
                    start_date,
                    end_date,
                    st.session_state.exclusions,
                    required_boards,
                    required_occupancies,
                    hotel_filter=hotel_options[hotel_filter],
                    city_filter=city_filter if city_filter != "All" else None,
                )

            st.session_state.gaps_df = gaps_df

            if len(gaps_df) == 0:
                st.success("🎉 No gaps found! All hotels have complete coverage.")
            else:
                # Summary metrics
                st.markdown("### 📊 Gap Summary")

                date_gaps = gaps_df.filter(pl.col("gap_type") == "date")
                board_gaps = gaps_df.filter(pl.col("gap_type") == "board")
                occ_gaps = gaps_df.filter(pl.col("gap_type") == "occupancy")

                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total Gaps", len(gaps_df))
                col2.metric("Date Gaps", len(date_gaps))
                col3.metric("Board Gaps", len(board_gaps))
                col4.metric("Occupancy Gaps", len(occ_gaps))

                hotels_with_gaps = gaps_df["hotel_id"].n_unique()
                total_hotels = daily_df["hotel_id"].n_unique()
                st.metric("Hotels with Gaps", f"{hotels_with_gaps} / {total_hotels}")

                st.markdown("---")

                # Gap tables by type
                st.markdown("### 🔴 Date Gaps (No Availability)")
                if len(date_gaps) > 0:
                    display_df = date_gaps.select([
                        "hotel_name", "city", "star_rating", "supplier_name",
                        "gap_start", "gap_end", "duration_days"
                    ]).with_columns([
                        pl.col("gap_start").dt.strftime("%d-%m-%Y").alias("gap_start"),
                        pl.col("gap_end").dt.strftime("%d-%m-%Y").alias("gap_end"),
                    ]).sort("duration_days", descending=True)
                    st.dataframe(display_df.to_pandas(), hide_index=True, use_container_width=True)
                else:
                    st.success("No date gaps found!")

                st.markdown("### 🟡 Board Gaps")
                if len(board_gaps) > 0:
                    display_df = board_gaps.select([
                        "hotel_name", "city", "star_rating", "supplier_name",
                        "detail", "gap_start", "gap_end", "duration_days"
                    ]).with_columns([
                        pl.col("gap_start").dt.strftime("%d-%m-%Y").alias("gap_start"),
                        pl.col("gap_end").dt.strftime("%d-%m-%Y").alias("gap_end"),
                    ]).sort("duration_days", descending=True)
                    st.dataframe(display_df.to_pandas(), hide_index=True, use_container_width=True)
                else:
                    st.success("No board gaps found!")

                st.markdown("### 🟠 Occupancy Gaps")
                if len(occ_gaps) > 0:
                    display_df = occ_gaps.select([
                        "hotel_name", "city", "star_rating", "supplier_name",
                        "detail", "gap_start", "gap_end", "duration_days"
                    ]).with_columns([
                        pl.col("gap_start").dt.strftime("%d-%m-%Y").alias("gap_start"),
                        pl.col("gap_end").dt.strftime("%d-%m-%Y").alias("gap_end"),
                    ]).sort("duration_days", descending=True)
                    st.dataframe(display_df.to_pandas(), hide_index=True, use_container_width=True)
                else:
                    st.success("No occupancy gaps found!")

        # Export
        st.markdown("---")
        st.markdown("### 📥 Export")

        if "gaps_df" in st.session_state and len(st.session_state.gaps_df) > 0:
            col_exp1, col_exp2 = st.columns(2)

            with col_exp1:
                # Simple CSV export (gap report only)
                export_df = st.session_state.gaps_df.with_columns([
                    pl.col("gap_start").dt.strftime("%d-%m-%Y").alias("gap_start"),
                    pl.col("gap_end").dt.strftime("%d-%m-%Y").alias("gap_end"),
                ])
                csv_buffer = io.StringIO()
                export_df.to_pandas().to_csv(csv_buffer, index=False)
                csv_data = csv_buffer.getvalue()

                st.download_button(
                    label="📥 Download Gaps (CSV)",
                    data=csv_data,
                    file_name=f"hotel_gaps_{date.today().strftime('%d-%m-%Y')}.csv",
                    mime="text/csv",
                )

            with col_exp2:
                # Enhanced Excel template with reference sheets
                from openpyxl import Workbook

                # Create workbook
                wb = Workbook()

                # Sheet 1: Gap Template - expanded by room type (one row per room type per gap)
                ws_gaps = wb.active
                ws_gaps.title = "Gaps_Template"

                # Headers for expanded template
                headers = [
                    "hotel_id", "organization_id", "hotel_name", "city", "star_rating",
                    "gap_type", "detail", "start_date", "end_date", "duration_days",
                    "supplier_name", "room_type_id", "room_name", "max_occupancy",
                    "supplier_id_to_fill", "occupancy", "weekday_rate", "weekend_rate",
                    "currency", "rate_type", "min_booking_days_in_advance", "num_of_rooms",
                    "included_meal_type_code"
                ]
                for col_idx, header in enumerate(headers, 1):
                    ws_gaps.cell(row=1, column=col_idx, value=header)

                # Build hotel -> room_types cache
                hotel_room_types_cache = {}
                unique_hotels = st.session_state.gaps_df.select(["hotel_id"]).unique()
                for hotel_row in unique_hotels.iter_rows(named=True):
                    hotel_id = hotel_row["hotel_id"]
                    hotel_room_types_cache[hotel_id] = load_room_types_for_hotel(hotel_id)

                # Expand gaps by room type
                row_idx = 2
                for gap_row in st.session_state.gaps_df.iter_rows(named=True):
                    hotel_id = gap_row["hotel_id"]
                    room_types = hotel_room_types_cache.get(hotel_id, [])

                    # Format dates
                    start_date_str = gap_row["gap_start"].strftime("%Y-%m-%d") if hasattr(gap_row["gap_start"], "strftime") else str(gap_row["gap_start"])
                    end_date_str = gap_row["gap_end"].strftime("%Y-%m-%d") if hasattr(gap_row["gap_end"], "strftime") else str(gap_row["gap_end"])

                    if not room_types:
                        # No room types - still create one row with empty room info
                        ws_gaps.cell(row=row_idx, column=1, value=hotel_id)
                        ws_gaps.cell(row=row_idx, column=2, value=gap_row.get("organization_id", ""))
                        ws_gaps.cell(row=row_idx, column=3, value=gap_row["hotel_name"])
                        ws_gaps.cell(row=row_idx, column=4, value=gap_row["city"])
                        ws_gaps.cell(row=row_idx, column=5, value=gap_row["star_rating"])
                        ws_gaps.cell(row=row_idx, column=6, value=gap_row["gap_type"])
                        ws_gaps.cell(row=row_idx, column=7, value=gap_row["detail"])
                        ws_gaps.cell(row=row_idx, column=8, value=start_date_str)
                        ws_gaps.cell(row=row_idx, column=9, value=end_date_str)
                        ws_gaps.cell(row=row_idx, column=10, value=gap_row["duration_days"])
                        ws_gaps.cell(row=row_idx, column=11, value=gap_row["supplier_name"])
                        ws_gaps.cell(row=row_idx, column=19, value="SAR")
                        ws_gaps.cell(row=row_idx, column=20, value="subject_to_availability")
                        row_idx += 1
                    else:
                        # Create one row per room type
                        for rt in room_types:
                            ws_gaps.cell(row=row_idx, column=1, value=hotel_id)
                            ws_gaps.cell(row=row_idx, column=2, value=gap_row.get("organization_id", ""))
                            ws_gaps.cell(row=row_idx, column=3, value=gap_row["hotel_name"])
                            ws_gaps.cell(row=row_idx, column=4, value=gap_row["city"])
                            ws_gaps.cell(row=row_idx, column=5, value=gap_row["star_rating"])
                            ws_gaps.cell(row=row_idx, column=6, value=gap_row["gap_type"])
                            ws_gaps.cell(row=row_idx, column=7, value=gap_row["detail"])
                            ws_gaps.cell(row=row_idx, column=8, value=start_date_str)
                            ws_gaps.cell(row=row_idx, column=9, value=end_date_str)
                            ws_gaps.cell(row=row_idx, column=10, value=gap_row["duration_days"])
                            ws_gaps.cell(row=row_idx, column=11, value=gap_row["supplier_name"])
                            ws_gaps.cell(row=row_idx, column=12, value=str(rt["id"]))  # room_type_id pre-filled
                            ws_gaps.cell(row=row_idx, column=13, value=rt["name"])  # room_name pre-filled
                            ws_gaps.cell(row=row_idx, column=14, value=rt["max_occupancy"])  # max_occupancy pre-filled
                            ws_gaps.cell(row=row_idx, column=19, value="SAR")
                            ws_gaps.cell(row=row_idx, column=20, value="subject_to_availability")
                            row_idx += 1

                # Sheet 2: Suppliers Reference
                ws_suppliers = wb.create_sheet("Suppliers")
                ws_suppliers.cell(row=1, column=1, value="supplier_id")
                ws_suppliers.cell(row=1, column=2, value="supplier_name")
                for row_idx, supplier in enumerate(all_suppliers, 2):
                    ws_suppliers.cell(row=row_idx, column=1, value=str(supplier["id"]))
                    ws_suppliers.cell(row=row_idx, column=2, value=supplier["name"])

                # Sheet 3: Room Types Reference
                ws_rooms = wb.create_sheet("Room_Types")
                ws_rooms.cell(row=1, column=1, value="hotel_id")
                ws_rooms.cell(row=1, column=2, value="hotel_name")
                ws_rooms.cell(row=1, column=3, value="room_type_id")
                ws_rooms.cell(row=1, column=4, value="room_name")
                ws_rooms.cell(row=1, column=5, value="max_occupancy")
                row_idx = 2
                # Get unique hotels from gaps
                unique_hotels = st.session_state.gaps_df.select(["hotel_id", "hotel_name"]).unique()
                for hotel_row in unique_hotels.iter_rows(named=True):
                    room_types = load_room_types_for_hotel(hotel_row["hotel_id"])
                    for rt in room_types:
                        ws_rooms.cell(row=row_idx, column=1, value=hotel_row["hotel_id"])
                        ws_rooms.cell(row=row_idx, column=2, value=hotel_row["hotel_name"])
                        ws_rooms.cell(row=row_idx, column=3, value=str(rt["id"]))
                        ws_rooms.cell(row=row_idx, column=4, value=rt["name"])
                        ws_rooms.cell(row=row_idx, column=5, value=rt["max_occupancy"])
                        row_idx += 1

                # Sheet 4: Meal Types Reference
                ws_meals = wb.create_sheet("Meal_Types")
                ws_meals.cell(row=1, column=1, value="meal_type_code")
                ws_meals.cell(row=1, column=2, value="meal_type_name")
                meal_types = load_all_meal_types()
                for row_idx, mt in enumerate(meal_types, 2):
                    ws_meals.cell(row=row_idx, column=1, value=mt["code"])
                    ws_meals.cell(row=row_idx, column=2, value=mt["name"])

                # Sheet 5: Occupancy Codes Reference
                ws_occ = wb.create_sheet("Occupancy_Codes")
                ws_occ.cell(row=1, column=1, value="occupancy_code")
                ws_occ.cell(row=1, column=2, value="occupancy_name")
                ws_occ.cell(row=1, column=3, value="capacity")
                for row_idx, (name, code) in enumerate(OCCUPANCY_CODES.items(), 2):
                    ws_occ.cell(row=row_idx, column=1, value=code)
                    ws_occ.cell(row=row_idx, column=2, value=name)
                    ws_occ.cell(row=row_idx, column=3, value=REQUIRED_OCCUPANCIES[name])

                # Save to bytes
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_data = excel_buffer.getvalue()

                st.download_button(
                    label="📥 Download Rate Template (Excel)",
                    data=excel_data,
                    file_name=f"gap_rate_template_{date.today().strftime('%d-%m-%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

            st.caption("Excel template includes reference sheets for suppliers, room types, meal types, and occupancy codes.")
        else:
            st.info("Generate a gap report first to enable export")

    # TAB 2: By Supplier
    with tab2:
        st.subheader("Gaps by Supplier")

        if "gaps_df" in st.session_state and len(st.session_state.gaps_df) > 0:
            supplier_summary = get_supplier_summary(st.session_state.gaps_df)

            if len(supplier_summary) > 0:
                st.dataframe(
                    supplier_summary.to_pandas(),
                    hide_index=True,
                    use_container_width=True,
                )

                # Chart
                fig = px.bar(
                    supplier_summary.to_pandas(),
                    x="supplier_name",
                    y="total_gap_days",
                    color="hotels_affected",
                    title="Total Gap Days by Supplier",
                    labels={
                        "supplier_name": "Supplier",
                        "total_gap_days": "Total Gap Days",
                        "hotels_affected": "Hotels",
                    },
                )
                st.plotly_chart(fig, use_container_width=True)

                # Detailed view per supplier
                st.markdown("---")
                st.markdown("### Detailed Gaps by Supplier")

                supplier_names = st.session_state.gaps_df["supplier_name"].unique().to_list()
                selected_supplier = st.selectbox("Select Supplier", supplier_names)

                if selected_supplier:
                    supplier_gaps = st.session_state.gaps_df.filter(
                        pl.col("supplier_name") == selected_supplier
                    )
                    display_df = supplier_gaps.select([
                        "hotel_name", "city", "star_rating", "gap_type",
                        "detail", "gap_start", "gap_end", "duration_days"
                    ]).with_columns([
                        pl.col("gap_start").dt.strftime("%d-%m-%Y").alias("gap_start"),
                        pl.col("gap_end").dt.strftime("%d-%m-%Y").alias("gap_end"),
                    ]).sort(["hotel_name", "gap_start"])

                    st.dataframe(display_df.to_pandas(), hide_index=True, use_container_width=True)

                    # Export supplier gaps
                    csv_buffer = io.StringIO()
                    display_df.to_pandas().to_csv(csv_buffer, index=False)
                    st.download_button(
                        label=f"📥 Download {selected_supplier} Gaps",
                        data=csv_buffer.getvalue(),
                        file_name=f"gaps_{selected_supplier.replace(' ', '_')}_{date.today().strftime('%d-%m-%Y')}.csv",
                        mime="text/csv",
                    )
        else:
            st.info("Generate a gap report first in the 'Gap Report' tab")

    # TAB 3: Summary
    with tab3:
        st.subheader("Coverage Summary")

        # Overall stats
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Hotels", daily_df["hotel_id"].n_unique())
        col2.metric("Total Suppliers", daily_df["supplier_name"].n_unique())
        col3.metric("Date Range", f"{start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}")

        # City breakdown
        st.markdown("### Hotels by City")
        city_stats = daily_df.group_by("city").agg([
            pl.col("hotel_id").n_unique().alias("hotels"),
        ])
        st.dataframe(city_stats.to_pandas(), hide_index=True)

        # Star rating breakdown
        st.markdown("### Hotels by Star Rating")
        star_stats = daily_df.group_by("star_rating").agg([
            pl.col("hotel_id").n_unique().alias("hotels"),
        ]).sort("star_rating")
        st.dataframe(star_stats.to_pandas(), hide_index=True)

        if "gaps_df" in st.session_state and len(st.session_state.gaps_df) > 0:
            st.markdown("---")
            st.markdown("### Gap Distribution")

            # Gap type pie chart
            gap_type_counts = st.session_state.gaps_df.group_by("gap_type").agg([
                pl.len().alias("count")
            ])

            fig = px.pie(
                gap_type_counts.to_pandas(),
                values="count",
                names="gap_type",
                title="Gaps by Type",
                color_discrete_sequence=["#EF553B", "#FECB52", "#FF7F0E"],
            )
            st.plotly_chart(fig, use_container_width=True)

            # Gaps by city
            city_gap_counts = st.session_state.gaps_df.group_by("city").agg([
                pl.len().alias("gaps"),
                pl.col("duration_days").sum().alias("total_days"),
            ])
            st.dataframe(city_gap_counts.to_pandas(), hide_index=True)

    # TAB 4: Visualizations
    with tab4:
        st.subheader("Gap Visualizations")

        if "gaps_df" not in st.session_state or len(st.session_state.gaps_df) == 0:
            st.info("Generate a gap report first in the 'Gap Report' tab to see visualizations")
        else:
            gaps_df = st.session_state.gaps_df

            # 1. Calendar Heatmap - Gap density by date
            st.markdown("### 📅 Gap Calendar Heatmap")
            st.caption("Shows the number of gaps per day across all hotels")

            # Expand gaps to daily and count gaps per date
            gap_dates = []
            for row in gaps_df.iter_rows(named=True):
                current = row["gap_start"]
                while current <= row["gap_end"]:
                    gap_dates.append({
                        "date": current,
                        "gap_type": row["gap_type"],
                        "hotel_id": row["hotel_id"],
                    })
                    current = current + timedelta(days=1)

            if gap_dates:
                gap_daily = pl.DataFrame(gap_dates)
                gap_counts = gap_daily.group_by("date").agg([
                    pl.len().alias("gap_count"),
                    pl.col("hotel_id").n_unique().alias("hotels_affected"),
                ]).sort("date")

                # Create calendar heatmap
                gap_counts_pd = gap_counts.to_pandas()
                gap_counts_pd["week"] = gap_counts_pd["date"].dt.isocalendar().week
                gap_counts_pd["weekday"] = gap_counts_pd["date"].dt.dayofweek
                gap_counts_pd["month"] = gap_counts_pd["date"].dt.strftime("%Y-%m")

                fig = px.density_heatmap(
                    gap_counts_pd,
                    x="date",
                    y="gap_count",
                    title="Gap Density Over Time",
                    labels={"date": "Date", "gap_count": "Number of Gaps"},
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)

                # Monthly summary heatmap
                st.markdown("### 📊 Monthly Gap Summary")
                monthly_gaps = gap_counts_pd.groupby("month").agg({
                    "gap_count": "sum",
                    "hotels_affected": "max"
                }).reset_index()

                fig2 = px.bar(
                    monthly_gaps,
                    x="month",
                    y="gap_count",
                    color="hotels_affected",
                    title="Total Gaps by Month",
                    labels={
                        "month": "Month",
                        "gap_count": "Total Gap Days",
                        "hotels_affected": "Max Hotels Affected"
                    },
                )
                fig2.update_layout(height=400)
                st.plotly_chart(fig2, use_container_width=True)

            # 2. Gap Timeline (Gantt-style)
            st.markdown("### 📈 Gap Timeline by Hotel")
            st.caption("Gantt chart showing gap periods for each hotel")

            # Filter options for timeline
            timeline_city = st.selectbox(
                "Filter by City",
                ["All"] + gaps_df["city"].unique().to_list(),
                key="timeline_city"
            )

            timeline_gap_type = st.selectbox(
                "Filter by Gap Type",
                ["All", "date", "board", "occupancy"],
                key="timeline_gap_type"
            )

            filtered_gaps = gaps_df
            if timeline_city != "All":
                filtered_gaps = filtered_gaps.filter(pl.col("city") == timeline_city)
            if timeline_gap_type != "All":
                filtered_gaps = filtered_gaps.filter(pl.col("gap_type") == timeline_gap_type)

            if len(filtered_gaps) > 0:
                # Limit to top 30 hotels for readability
                top_hotels = filtered_gaps.group_by("hotel_name").agg([
                    pl.col("duration_days").sum().alias("total_days")
                ]).sort("total_days", descending=True).head(30)["hotel_name"].to_list()

                timeline_data = filtered_gaps.filter(
                    pl.col("hotel_name").is_in(top_hotels)
                ).to_pandas()

                if len(timeline_data) > 0:
                    # Create Gantt chart
                    fig3 = px.timeline(
                        timeline_data,
                        x_start="gap_start",
                        x_end="gap_end",
                        y="hotel_name",
                        color="gap_type",
                        hover_data=["supplier_name", "detail", "duration_days"],
                        title=f"Gap Timeline (Top {len(top_hotels)} Hotels by Gap Days)",
                        color_discrete_map={
                            "date": "#EF553B",
                            "board": "#FECB52",
                            "occupancy": "#FF7F0E"
                        },
                    )
                    fig3.update_layout(
                        height=max(400, len(top_hotels) * 25),
                        yaxis_title="Hotel",
                        xaxis_title="Date",
                    )
                    st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("No gaps match the selected filters")

            # 3. Hotel Coverage Heatmap
            st.markdown("### 🏨 Hotel Coverage Matrix")
            st.caption("Shows which hotels have gaps in which months")

            # Create hotel x month matrix
            hotel_month_gaps = []
            for row in gaps_df.iter_rows(named=True):
                current = row["gap_start"]
                while current <= row["gap_end"]:
                    hotel_month_gaps.append({
                        "hotel_name": row["hotel_name"],
                        "month": current.strftime("%Y-%m"),
                        "gap_type": row["gap_type"],
                    })
                    current = current + timedelta(days=1)

            if hotel_month_gaps:
                hm_df = pl.DataFrame(hotel_month_gaps)
                hm_summary = hm_df.group_by(["hotel_name", "month"]).agg([
                    pl.len().alias("gap_days")
                ]).to_pandas()

                # Pivot for heatmap
                hm_pivot = hm_summary.pivot(
                    index="hotel_name",
                    columns="month",
                    values="gap_days"
                ).fillna(0)

                # Limit to top 20 hotels
                hotel_totals = hm_pivot.sum(axis=1).sort_values(ascending=False)
                top_20 = hotel_totals.head(20).index.tolist()
                hm_pivot_top = hm_pivot.loc[top_20]

                fig4 = go.Figure(data=go.Heatmap(
                    z=hm_pivot_top.values,
                    x=hm_pivot_top.columns.tolist(),
                    y=hm_pivot_top.index.tolist(),
                    colorscale="Reds",
                    hoverongaps=False,
                    hovertemplate="Hotel: %{y}<br>Month: %{x}<br>Gap Days: %{z}<extra></extra>",
                ))
                fig4.update_layout(
                    title="Gap Days by Hotel and Month (Top 20 Hotels)",
                    xaxis_title="Month",
                    yaxis_title="Hotel",
                    height=max(400, len(top_20) * 25),
                )
                st.plotly_chart(fig4, use_container_width=True)

            # 4. Gap Type Distribution Over Time
            st.markdown("### 📉 Gap Types Over Time")

            if gap_dates:
                gap_type_time = pl.DataFrame(gap_dates).group_by(["date", "gap_type"]).agg([
                    pl.len().alias("count")
                ]).sort("date").to_pandas()

                fig5 = px.area(
                    gap_type_time,
                    x="date",
                    y="count",
                    color="gap_type",
                    title="Gap Types Distribution Over Time",
                    labels={"date": "Date", "count": "Number of Gaps", "gap_type": "Gap Type"},
                    color_discrete_map={
                        "date": "#EF553B",
                        "board": "#FECB52",
                        "occupancy": "#FF7F0E"
                    },
                )
                fig5.update_layout(height=400)
                st.plotly_chart(fig5, use_container_width=True)

    # TAB 5: Fill Gaps
    with tab5:
        st.subheader("Create Rates to Fill Gaps")

        # Check Hasura connection
        hasura_connected = test_hasura_connection()
        if not hasura_connected:
            st.warning("Hasura GraphQL not configured. Add HASURA_GRAPHQL_URL and HASURA_ADMIN_SECRET to .env file to enable rate creation.")

        if "gaps_df" not in st.session_state or len(st.session_state.gaps_df) == 0:
            st.info("Generate a gap report first in the 'Gap Report' tab to fill gaps.")
        else:
            gaps_df = st.session_state.gaps_df

            st.markdown("### 1. Select Gap to Fill")

            # Create display options for gap selection
            gap_options = []
            for i, row in enumerate(gaps_df.iter_rows(named=True)):
                gap_start_str = row["gap_start"].strftime("%d-%m-%Y") if hasattr(row["gap_start"], "strftime") else str(row["gap_start"])
                gap_end_str = row["gap_end"].strftime("%d-%m-%Y") if hasattr(row["gap_end"], "strftime") else str(row["gap_end"])
                label = f"{row['hotel_name']} | {row['gap_type']} | {row['detail']} | {gap_start_str} to {gap_end_str}"
                gap_options.append(label)

            selected_gap_idx = st.selectbox(
                "Select a gap to fill",
                options=range(len(gap_options)),
                format_func=lambda i: gap_options[i],
                key="fill_gap_select"
            )

            # Get selected gap data
            selected_gap = gaps_df[selected_gap_idx].to_dicts()[0]

            st.markdown("### 2. Gap Details (Read-Only)")
            col1, col2, col3 = st.columns(3)
            col1.text_input("Hotel", selected_gap["hotel_name"], disabled=True, key="gap_hotel")
            col2.text_input("City", selected_gap["city"], disabled=True, key="gap_city")
            col3.text_input("Star Rating", str(selected_gap["star_rating"]), disabled=True, key="gap_stars")

            col1, col2, col3 = st.columns(3)
            col1.text_input("Gap Type", selected_gap["gap_type"], disabled=True, key="gap_type_display")
            col2.text_input("Gap Start", selected_gap["gap_start"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_start"], "strftime") else str(selected_gap["gap_start"]), disabled=True, key="gap_start_display")
            col3.text_input("Gap End", selected_gap["gap_end"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_end"], "strftime") else str(selected_gap["gap_end"]), disabled=True, key="gap_end_display")

            st.text_input("Detail", selected_gap["detail"], disabled=True, key="gap_detail_display")

            st.markdown("### 3. Rate Details (Fill In)")

            # Supplier dropdown
            supplier_options = {s["name"]: str(s["id"]) for s in all_suppliers}
            selected_supplier_name = st.selectbox(
                "Supplier",
                list(supplier_options.keys()),
                key="rate_supplier"
            )
            selected_supplier_id = supplier_options[selected_supplier_name]

            # Room type dropdown (filtered by hotel)
            hotel_id = selected_gap["hotel_id"]
            room_types = load_room_types_for_hotel(hotel_id)

            if not room_types:
                st.warning(f"No room types found for this hotel. Please add room types first.")
                room_type_options = {}
            else:
                room_type_options = {f"{rt['name']} (max {rt['max_occupancy']})": str(rt["id"]) for rt in room_types}

            selected_room_type_name = st.selectbox(
                "Room Type",
                list(room_type_options.keys()) if room_type_options else ["No room types available"],
                key="rate_room_type",
                disabled=not room_type_options
            )
            selected_room_type_id = room_type_options.get(selected_room_type_name, "")

            # Occupancy
            occupancy = st.selectbox(
                "Occupancy",
                list(OCCUPANCY_CODES.keys()),
                key="rate_occupancy",
                format_func=lambda x: f"{x} ({OCCUPANCY_CODES[x]})"
            )
            occupancy_code = OCCUPANCY_CODES[occupancy]

            # Rates
            col1, col2 = st.columns(2)
            with col1:
                weekday_rate = st.number_input("Weekday Rate", min_value=0.0, step=10.0, key="rate_weekday")
            with col2:
                weekend_rate = st.number_input("Weekend Rate", min_value=0.0, step=10.0, key="rate_weekend")

            # Currency and rate type
            col1, col2 = st.columns(2)
            with col1:
                currency = st.selectbox("Currency", ["SAR", "USD", "EUR"], key="rate_currency")
            with col2:
                rate_type = st.selectbox(
                    "Rate Type",
                    ["subject_to_availability", "guaranteed"],
                    key="rate_type_select"
                )

            # Meal type
            meal_types = load_all_meal_types()
            meal_type_options = {mt["name"]: mt["code"] for mt in meal_types}
            selected_meal_name = st.selectbox(
                "Included Meal",
                list(meal_type_options.keys()),
                key="rate_meal_type"
            )
            selected_meal_code = meal_type_options[selected_meal_name]

            # Optional fields
            col1, col2 = st.columns(2)
            with col1:
                min_booking_days = st.number_input(
                    "Min Booking Days in Advance",
                    min_value=0,
                    value=0,
                    key="rate_min_booking"
                )
            with col2:
                num_rooms = st.number_input(
                    "Number of Rooms",
                    min_value=1,
                    value=1,
                    key="rate_num_rooms"
                )

            st.markdown("### 4. Submit")

            # Show what will be created
            with st.expander("Preview Rate Data"):
                st.json({
                    "organization_id": selected_gap.get("organization_id", "N/A"),
                    "hotel_id": hotel_id,
                    "room_type_id": selected_room_type_id,
                    "supplier_id": selected_supplier_id,
                    "start_date": selected_gap["gap_start"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_start"], "strftime") else str(selected_gap["gap_start"]),
                    "end_date": selected_gap["gap_end"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_end"], "strftime") else str(selected_gap["gap_end"]),
                    "occupancy": occupancy_code,
                    "weekday_rate": weekday_rate,
                    "weekend_rate": weekend_rate,
                    "currency": currency,
                    "rate_type": rate_type,
                    "included_meal_type_code": selected_meal_code,
                    "min_booking_days_in_advance": min_booking_days if min_booking_days > 0 else None,
                    "num_of_rooms": num_rooms,
                    "status": "pending_approval"
                })

            # Submit button
            submit_disabled = not hasura_connected or not room_type_options or weekday_rate <= 0 or weekend_rate <= 0

            if st.button("Create Rate", type="primary", disabled=submit_disabled, key="submit_rate"):
                # Prepare dates
                start_date_str = selected_gap["gap_start"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_start"], "strftime") else str(selected_gap["gap_start"])
                end_date_str = selected_gap["gap_end"].strftime("%Y-%m-%d") if hasattr(selected_gap["gap_end"], "strftime") else str(selected_gap["gap_end"])

                with st.spinner("Creating rate..."):
                    result = insert_hotel_rate(
                        organization_id=selected_gap.get("organization_id", ""),
                        hotel_id=hotel_id,
                        room_type_id=selected_room_type_id,
                        supplier_id=selected_supplier_id,
                        start_date=start_date_str,
                        end_date=end_date_str,
                        occupancy=occupancy_code,
                        weekday_rate=weekday_rate,
                        weekend_rate=weekend_rate,
                        currency=currency,
                        rate_type=rate_type,
                        included_meal_type_code=selected_meal_code,
                        min_booking_days_in_advance=min_booking_days if min_booking_days > 0 else None,
                        num_of_rooms=num_rooms,
                    )

                if "errors" in result:
                    error_msg = result["errors"][0].get("message", "Unknown error")
                    if "unique constraint" in error_msg.lower():
                        st.error("A rate with these exact dates, room type, supplier, and occupancy already exists.")
                    else:
                        st.error(f"Error creating rate: {error_msg}")
                else:
                    rate_id = result.get("data", {}).get("insert_hotel_rates_one", {}).get("id", "N/A")
                    st.success(f"Rate created successfully! ID: {rate_id}")
                    st.info("Note: Rate status is 'pending_approval'. Refresh data to see updated gaps.")

                    # Clear rate cache
                    if st.button("Refresh Data", key="refresh_after_create"):
                        load_all_rates.clear()
                        st.rerun()

            if submit_disabled and hasura_connected:
                if not room_type_options:
                    st.caption("Cannot submit: No room types available for this hotel.")
                elif weekday_rate <= 0 or weekend_rate <= 0:
                    st.caption("Cannot submit: Rates must be greater than 0.")


def main():
    """Main application entry point."""
    # Load authentication config
    config = load_auth_config()

    # Create authenticator
    authenticator = stauth.Authenticate(
        config["credentials"],
        config["cookie"]["name"],
        config["cookie"]["key"],
        config["cookie"]["expiry_days"],
    )

    # Login
    authenticator.login(location="main")

    if st.session_state.get("authentication_status"):
        # Show logout in sidebar
        authenticator.logout("Logout", "sidebar")
        st.sidebar.write(f"Welcome, {st.session_state.get('name')}")

        # Show main dashboard
        main_dashboard()

    elif st.session_state.get("authentication_status") == False:
        st.error("Username/password is incorrect")

    elif st.session_state.get("authentication_status") is None:
        st.warning("Please enter your username and password")


if __name__ == "__main__":
    main()
